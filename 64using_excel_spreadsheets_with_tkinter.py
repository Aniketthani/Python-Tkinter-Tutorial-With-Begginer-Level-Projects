from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


def get_a():
    a=""
    for cell in column_a:
        a=a+str(cell.value) + "\n"
    label_a.config(text=a)

def get_b():
    b=""
    for cell in column_b:
        b=b+str(cell.value) + "\n"
    label_b.config(text=b)
    

root=Tk()
root.title("Using Excel Spreadsheets with tkinter by the help of openpyxl module of python")
root.geometry("500x500")


#create a workbook instance
wb=Workbook()

# load existing workbook
wb=load_workbook("test.xlsx")

#create active worksheet
ws=wb.active

#create variable and store columns
column_a=ws["A"]
column_b=ws["B"]

#print(column_a) 
# output :(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>, <Cell 'Sheet1'.A5>, <Cell 'Sheet1'.A6>, <Cell 'Sheet1'.A7>, <Cell 'Sheet1'.A8>, <Cell 'Sheet1'.A9>)
#print(column_b) 
# output :(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>, <Cell 'Sheet1'.B8>, <Cell 'Sheet1'.B9>)



#add a record to the xl file and save it as a new file
ws["A9"]="I"
ws["B9"]=500
#save the file as new file
wb.save("test2.xlsx")


b_a=Button(root,text="Get Column A",command=get_a)
b_a.pack(pady=20)

label_a = Label(root, text="")
label_a.pack()



b_b=Button(root,text="Get Column B",command=get_b)
b_b.pack(pady=20)

label_b = Label(root, text="")
label_b.pack()





root.mainloop()
