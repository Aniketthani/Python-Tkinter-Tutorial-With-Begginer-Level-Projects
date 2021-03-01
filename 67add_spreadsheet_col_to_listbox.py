from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


def insert(col):
    l1.delete(0,END)
    for cell in col:
        l1.insert(END,str(cell.value))
root=Tk()
root.geometry("700x700")


#create workbook instance
wb=Workbook()

#load spreadsheet or existing workbook

wb=load_workbook("test.xlsx")

#grab active worksheet

ws=wb.active

#columns
column_a=ws['A']
column_b=ws["B"]

#listbox
l1=Listbox(root,width=30)
l1.pack(pady=20)

insert_col_a_button=Button(root,text="insert col a to listbox",command=lambda: insert(column_a))
insert_col_b_button=Button(root,text="insert col b to listbox",command=lambda : insert(column_b))

insert_col_a_button.pack(pady=10)
insert_col_b_button.pack(pady=10)



root.mainloop()